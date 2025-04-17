#!/usr/bin/env python3
"""
Excel file consolidation module for the ops-sim project.
This module provides functionality for consolidating Excel files into a master file.
"""

# Standard library imports
import os
import re
import traceback
from pathlib import Path
from typing import Dict, List, Optional, Union

# Third-party imports
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

# Local imports
import config


def copy_formatting(source_sheet: Worksheet, target_sheet: Worksheet) -> None:
    """Copy formatting from source sheet to target sheet.
    
    This includes:
    - Column widths
    - Font styles
    - Cell fills
    - Alignment
    - Borders
    - Number formats
    
    Args:
        source_sheet: Worksheet to copy formatting from
        target_sheet: Worksheet to copy formatting to
    """
    # Copy exact column widths
    for col in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width

    # Copy cell formatting from the header row only
    for cell in next(source_sheet.iter_rows()):
        if cell.has_style:
            target_cell = target_sheet.cell(row=1, column=cell.column)
            
            # Copy font
            target_cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                vertAlign=cell.font.vertAlign,
                underline=cell.font.underline,
                strike=cell.font.strike,
                color=cell.font.color
            )
            
            # Copy fill
            target_cell.fill = PatternFill(
                start_color=cell.fill.start_color,
                end_color=cell.fill.end_color,
                fill_type=cell.fill.fill_type
            )
            
            # Copy alignment
            target_cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                text_rotation=cell.alignment.text_rotation,
                wrap_text=cell.alignment.wrap_text,
                shrink_to_fit=cell.alignment.shrink_to_fit,
                indent=cell.alignment.indent
            )
            
            # Copy borders
            target_cell.border = Border(
                left=Side(style=cell.border.left.style, color=cell.border.left.color),
                right=Side(style=cell.border.right.style, color=cell.border.right.color),
                top=Side(style=cell.border.top.style, color=cell.border.top.color),
                bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color)
            )
            
            # Copy number format
            target_cell.number_format = cell.number_format


def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Clean column names by removing 'Unnamed:' prefixes.
    
    Args:
        df: DataFrame with columns to clean
        
    Returns:
        DataFrame with cleaned column names
    """
    new_columns = {col: '' if col.startswith('Unnamed:') else col for col in df.columns}
    return df.rename(columns=new_columns)


def get_day_number(filename: Union[str, Path]) -> int:
    """Extract the day number from a filename.
    
    Args:
        filename: Name of the file to extract day number from
        
    Returns:
        Day number as integer, or 0 if not found
    """
    match = re.search(r'Day (\d+)', str(filename))
    return int(match.group(1)) if match else 0


def get_latest_excel_file() -> Optional[Path]:
    """Get the latest Excel file from the data folder.
    
    Returns:
        Path to the latest Excel file, or None if no files found
        
    Raises:
        FileNotFoundError: If the data folder doesn't exist
    """
    if not config.DATA_FOLDER_PATH.exists():
        raise FileNotFoundError(f"Data folder not found: {config.DATA_FOLDER_PATH}")
    
    # Get all Excel files excluding master and temp files
    excel_files = list(config.DATA_FOLDER_PATH.glob("*.xlsx"))
    excel_files = [f for f in excel_files if f.name != "Master.xlsx" and not f.name.startswith("~$")]
    
    if not excel_files:
        return None
    
    # Find latest file by day number in filename
    return max(excel_files, key=get_day_number)


def append_to_master() -> None:
    """Append data from the latest Excel file to the master file.
    
    This function:
    1. Finds the latest Excel file in the data folder
    2. Reads all sheets from the latest file
    3. Updates the master file with the new data
    4. Preserves formatting from the source file
    
    Raises:
        FileNotFoundError: If no Excel files are found
        ValueError: If the latest file cannot be read
        IOError: If there's an error writing to the master file
    """
    # Get the latest Excel file
    latest_file = get_latest_excel_file()
    if latest_file is None:
        raise FileNotFoundError("No Excel files found in directory")
    
    print(f"Processing latest file: {latest_file.name}")
    
    try:
        # Read all sheets from the latest file
        new_data_dict = pd.read_excel(
            latest_file,
            engine='openpyxl',
            sheet_name=None,
            keep_default_na=False,
            na_filter=False
        )
        
        # Load the source workbook for formatting
        source_wb = load_workbook(latest_file)
        
        # Create or update master file
        with pd.ExcelWriter(config.MASTER_FILE, engine='openpyxl') as writer:
            for sheet_name, new_data in new_data_dict.items():
                if sheet_name.endswith('-Graphs'):
                    print(f"Skipping graph sheet '{sheet_name}'")
                    continue
                
                print(f"\nProcessing sheet: {sheet_name}")
                print(f"New data rows: {len(new_data)}")
                
                # Clean the column names
                new_data = clean_column_names(new_data)
                
                # Write the new data to Excel
                new_data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Copy formatting from source
                if sheet_name in source_wb.sheetnames:
                    source_sheet = source_wb[sheet_name]
                    target_sheet = writer.sheets[sheet_name]
                    copy_formatting(source_sheet, target_sheet)
        
        print("\nMaster file updated successfully!")
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        print("Error details:")
        print(traceback.format_exc())
        raise


def main() -> None:
    """Main entry point for the script."""
    try:
        append_to_master()
    except Exception as e:
        print(f"Error in append: {str(e)}")
        raise


if __name__ == "__main__":
    main()
